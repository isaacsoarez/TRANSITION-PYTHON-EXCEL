import openpyxl, PatternFill

openpyxl.Workbook()

book = openpyxl.Workbook()
print(book.sheetnames)

book.create_sheet("frutas")
print(book.sheetnames)

book.create_sheet("bebidas")
print(book.sheetnames)

book["frutas"]

frutas_page = book["frutas"]

frutas_page.append(["Frutas", "Quantidade", "Preço"])
frutas_page.append(["Melancia", "12", "R$11,99"])
frutas_page.append(["Maçã", "07", "R$7,89"])
frutas_page.append(["Carambola", "23", "R$3,90"])

book["bebidas"]

bebidas_page = book["bebidas"]

bebidas_page.append(["Bebidas", "Quantidade", "Preço"])
bebidas_page.append(["Coca-Cola", "25", "R$9,89"])
bebidas_page.append(["Fanta Laranja", "17", "R$7,90"])
bebidas_page.append(["Chá Matte", "13", "R$5,60"])



def colorizar_celulas("Panilha de Compras.xlsx", "B2", "0000FF"):
    planilha['B2'].fill = PatternFill(start_color='0000FF', end_color='0000FF')

arquivo_excel = openpyxl.load_workbook('Panilha de Compras.xlsx')

# Escolhendo a planilha onde as células serão coloridas
planilha = arquivo_excel.active

# Colorindo a célula
colorizar_celulas("Panilha de Compras.xlsx', 'B2', '0000FF")


book.save("Panilha de Compras.xlsx")


