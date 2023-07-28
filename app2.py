import openpyxl

openpyxl.load_workbook('Panilha de Compras.xlsx')

book = openpyxl.load_workbook('Panilha de Compras.xlsx')

book["frutas"]

frutas_page = book["frutas"]

for rows in frutas_page.iter_rows(min_row=2, max_row=4):
    for cell in rows:
        print(cell.value)


for rows in frutas_page.iter_rows(min_row=2, max_row=4):
    print(rows[0].value, rows[1].value, rows[2].value)

for rows in frutas_page.iter_rows(min_row=2, max_row=4):
    for cell in rows:
        if cell.value == "Banana":
            cell.value = "Frutas 01"

book.save("planilha de Compras.xlsx")

book.save("planilha de Compras v2.xlsx")

#----------------------------------------------------------
book["bebidas"]

bebidas_page = book["bebidas"]

for rows in bebidas_page.iter_rows(min_row=2, max_row=4):
    for cell in rows:
        print(cell.value)


for rows in bebidas_page.iter_rows(min_row=2, max_row=4):
    print(rows[0].value, rows[1].value, rows[2].value)

    for rows in bebidas_page.iter_rows(min_row=2, max_row=4):
        for cell in rows:
            if cell.value == "Coca-Cola":
                cell.value = "bebidas 01"

    book.save("planilha de Compras.xlsx")

    book.save("planilha de Compras v2.xlsx")

